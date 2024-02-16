from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


url_path = './ojk/2.1 s.d 2.9 Tabel Keuangan Asuransi Umum dan Reasuransi 2022.xlsx'
pre_title = 'REKAPITULASI'
db_name='ASURANSI_UMUM'
def load():
    workbook = load_workbook(url_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_title = sheet.title.strip()
        #title = sheet_title.split(maxsplit=1)
        tablename=sheet_title.replace(' ','').replace('.','')
        tablename = ''.join([i for i in tablename if not i.isdigit()])
        print(tablename)
        
        if tablename=='Neraca':
            load_neraca(sheet_title,tablename)
        elif tablename=='Investasi':
            load_investasi(sheet_title,tablename)
        elif tablename=='NonInvestasi':
            load_non_investasi(sheet_title,tablename)    
        elif tablename=='HasilInvestasi':
            load_hasil_investasi(sheet_title,tablename)
        elif tablename=='StatNeraca':
            load_stat_neraca(sheet_title,tablename)   
        elif 'LABARUGI' in tablename.upper():
            load_rl(sheet_title,tablename)      
        elif tablename=='StatLR':
            load_stat_lr(sheet_title,'STATINCOME')  
        elif 'ASET' in tablename.upper():
            load_StatALE(sheet_title,'StatALE')
        elif 'LR-' in tablename.upper():
            load_StatRL(sheet_title,'StatRL')    
        elif 'INVDANCAD-' in tablename.upper():
            load_INVCAD(sheet_title,'INVCAD')
def load_INVCAD(sheet_name,table_name):
    import pandas as pd
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    columns = getcolumns(df,9,1)
    df= getdata(df,13,columns,1)
    columns=['REPORT_DATE','NAME OF COMPANIES','INVESTMENTS','TECHNICAL RESERVES','RATIO(TECHNICAL_RESERVES/INVESTMENTS)']
    df_new =pd.DataFrame(columns=columns) 
    df_new['NAME OF COMPANIES']=df['NAME OF COMPANIES']
    for year in range(2018,2023):
        date = getLastDateYear(year)
        df_new['REPORT_DATE']=date
        colYear = '{}'.format(year)
        df_new['INVESTMENTS']=df[year].iloc[:,0]
        df_new['TECHNICAL RESERVES']=df[year].iloc[:,1]
        df_new['RATIO(TECHNICAL_RESERVES/INVESTMENTS)']=df[year].iloc[:,2]
        df_new=df_new.assign(PRICE_INDEX=1000000)
        dump_to_db(table_name.upper(),df_new,date)
                  
def load_StatRL(sheet_name,table_name):
    import pandas as pd
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    columns = getcolumns(df,9,1)
    df= getdata(df,14,columns,1)
    columns=['REPORT_DATE','NAME OF COMPANIES','INCOMES','EXPENSES','PROFIT(LOSS)']
    df_new =pd.DataFrame(columns=columns) 
    df_new['NAME OF COMPANIES']=df['NAME OF COMPANIES']
    for year in range(2018,2023):
        date = getLastDateYear(year)
        df_new['REPORT_DATE']=date
        colYear = '{}'.format(year)
        df_new['INCOMES']=df[year].iloc[:,0]
        df_new['EXPENSES']=df[year].iloc[:,1]
        df_new['PROFIT(LOSS)']=df[year].iloc[:,2]
        df_new=df_new.assign(PRICE_INDEX=1000000)
        dump_to_db(table_name.upper(),df_new,date)

def load_StatALE(sheet_name,table_name):
    import pandas as pd
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    columns = getcolumns(df,9,1)
    df= getdata(df,11,columns,1)
    columns=['REPORT_DATE','NAME OF COMPANIES','ASSETS','LIABILITIES','EQUITIES']
    df_new =pd.DataFrame(columns=columns) 
    df_new['NAME OF COMPANIES']=df['NAME OF COMPANIES']
    for year in range(2018,2023):
        date = getLastDateYear(year)
        df_new['REPORT_DATE']=date
        colYear = '{}'.format(year)
        df_new['ASSETS']=df[year].iloc[:,0]
        df_new['LIABILITIES']=df[year].iloc[:,1]
        df_new['EQUITIES']=df[year].iloc[:,2]
        df_new=df_new.assign(PRICE_INDEX=1000000)
        dump_to_db(table_name.upper(),df_new,date)

def load_rl(sheet_name,table_name):
    import numpy as np
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    date = getReportDate(df,3)
    columns = getcolumns(df,8,1)
    columns=np.insert(columns,0,'NAME OF COMPANIES')
    df= getdata(df,13,columns,2)
    df.insert(0,'REPORT_DATE',date)
    df=df.assign(PRICE_INDEX=1000000)
    dump_to_db(table_name.upper(),df,date)

def load_stat_lr(sheet_name,table_name):
    import pandas as pd
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    columns =['INCOME','2018','2019','2020','2021','2022']
    df= getdata(df,9,columns,3)
    print(df)
    columns=['REPORT_DATE','INCOME','VALUES']
    df_new =pd.DataFrame(columns=columns) 
    df_new['INCOME']=df['INCOME']
    for year in range(2018,2023):
        date = getLastDateYear(year)
        df_new['REPORT_DATE']=date
        colYear = '{}'.format(year)
        df_new['VALUES']=df[colYear]
        df_new=df_new.assign(PRICE_INDEX=1000000)
        dump_to_db(table_name.upper(),df_new,date)

def load_stat_neraca(sheet_name,table_name):
    import pandas as pd
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    columns =['ASSETS','2018','2019','2020','2021','2022']
    df= getdata(df,9,columns,2)
    columns=['REPORT_DATE','ASSETS','VALUES']
    df_new =pd.DataFrame(columns=columns) 
    df_new['ASSETS']=df['ASSETS']
    for year in range(2018,2023):
        date = getLastDateYear(year)
        df_new['REPORT_DATE']=date
        colYear = '{}'.format(year)
        df_new['VALUES']=df[colYear]
        df_new=df_new.assign(PRICE_INDEX=1000000)
        dump_to_db(table_name.upper(),df_new,date)
def load_hasil_investasi(sheet_name,table_name):
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    date = getReportDate(df,3)
    columns = getcolumns(df,7,1)
    df= getdata(df,12,columns,1)
    df.insert(0,'REPORT_DATE',date)
    df=df.assign(PRICE_INDEX=1000000)
    dump_to_db(table_name.upper(),df,date)

def load_investasi(sheet_name,table_name):
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    date = getReportDate(df,3)
    columns = getcolumns(df,8)
    df= getdata(df,13,columns)
    df.insert(0,'REPORT_DATE',date)
    df=df.assign(PRICE_INDEX=1000000)
    dump_to_db(table_name.upper(),df,date)

def load_non_investasi(sheet_name,table_name):
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    date = getReportDate(df,3)
    columns = getcolumnsMultiRows(df,10,1)
    df= getdata(df,16,columns,1)
    df.insert(0,'REPORT_DATE',date)
    df=df.assign(PRICE_INDEX=1000000)
    dump_to_db(table_name.upper(),df,date)

def load_neraca(sheet_name,table_name):
    df = pd.read_excel(url_path,sheet_name=sheet_name)
    date = getReportDate(df,3)
    columns = getcolumnssplit(df,7)
    df= getdata(df,12,columns)
    df.insert(0,'REPORT_DATE',date)
    df=df.assign(PRICE_INDEX=1000000)
    dump_to_db(table_name.upper(),df,date)

def getReportDate(df,start):
    import datetime
    year = int(df.iloc[start].dropna().values[0].split()[-1])
    last_day_of_year = datetime.date.max.replace(year = year)
    return last_day_of_year

def getLastDateYear(year):
    import datetime
    last_day_of_year = datetime.date.max.replace(year = year)
    return last_day_of_year

def getFirstDateYear(year):
    import datetime
    first_day_of_year = datetime.date.min.replace(year=year)
    return first_day_of_year    

def getcolumnssplit(df,start):
    df=df.iloc[start].dropna()
    df = df.str.split('\n', expand=True)
    df.columns = ['IN','EN']
    columns = df['EN'].dropna().values
    return columns

def getcolumns(df,start_row,start_col=2):
    columns = df.iloc[start_row,start_col:].dropna().values
    return columns

def getcolumnsMultiRows(df,start_row,start_col=2):
    df1=df.iloc[start_row,start_col:]
    df2=df.iloc[start_row+1,start_col:]
    dfs = [df1,df2]
    df = pd.concat(dfs,axis=1)
    df.columns = ['col1','col2']
    df['columns'] = df['col1'].map(str) + ' ' + df['col2'].fillna('').map(str)
    columns = df['columns'].str.strip().dropna().values
    return columns

def getdata(df,start_row,columns,start_col=2):
    df = df[start_row:]
    df = df.iloc[:,start_col:]
    df.columns = columns
    df=df.dropna()
    df=df[df[columns[0]].str.contains( "JUMLAH" )==False]
    return df

def dump_to_db(table_name,df,report_date):
    from sqlalchemy import create_engine
    from sqlalchemy.exc import SQLAlchemyError
    db = create_engine('sqlite:///{}.db'.format(db_name))
    # drop 'records table'
    try:
        with db.connect() as conn:
            query="DELETE FROM {} WHERE REPORT_DATE='{}'".format(table_name,report_date)
            r_set=conn.exec_driver_sql(query)
    except SQLAlchemyError as e:
        error = str(e)
        print(error)
    else:
        print("No of Records deleted : ",r_set.rowcount)
    finally:
        df.to_sql(table_name, db, if_exists='append')

def get_data(dbname,tablename):
    import pandas as pd
    import sqlite3
    # Read sqlite query results into a pandas DataFrame
    con = sqlite3.connect(dbname)
    df = pd.read_sql_query('SELECT * from {}'.format(tablename), con)
    return df

def plotme(df,x,y,chart):
    import pandas as pd
    import matplotlib.pyplot as plt
    df.plot(x=x, y=y,
    kind=chart, figsize=(10, 10))
    plt.show()
    
def plotmultiseries(df,x,y,hue):
    import pandas as pd
    import matplotlib.pyplot as plt
    import seaborn as sns
    sns.lineplot(x=x, y=y, data=df, hue=hue)
    sns.set(style='dark',)
    plt.show()
    
      
    
    
if __name__=="__main__":
    load()